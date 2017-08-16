# encoding: utf-8

#-------------------
# built in libs
#-------------------
import os
import sys
import uuid
import html

#-------------------
# 3rd party libs
#-------------------
# http://pypi.python.org/pypi/xlrd
import openpyxl

#-------------------
# my scripts
#-------------------
from xls2expressionmap import constants
from xls2expressionmap.expressionmap import template
from xls2expressionmap.xlsx import xlsutil

def createUUID():
    return uuid.uuid4().fields[ 0 ]

def createUUIDList( listSize ):
    ret = [0] * listSize
    for i in range( listSize ):
          ret[ i ] = uuid.uuid4().fields[ 0 ]

    return ret

def float2int( v, defaultValue = 0 ):
    if( isinstance( v, float ) ):
        return int( v )
    else:
        return defaultValue

def getGroupValue( sheet, row ):
    group = xlsutil.getCellFromColmnName( sheet, row, "Group" )

    # Since v0.0.5: Colmn "Group" check (+backward compatibility)
    if( group != None and hasattr( group, "value" ) ):
        # float to int saferty
        group   = float2int( group.value ) - 1

        if( group < 0 ):
            group = 0
    else:
        group = 0

    return group

def genArticulation( sheet ):
    rowLength = sheet.max_row

    ret = template.ARTICULATION_HEADER

    for row in range( 2, rowLength ):

        name = xlsutil.getCellFromColmnName( sheet, row, "Articulation" )

        if len( name ) == 0:
            break

        name = name.strip()

        artiType = xlsutil.getCellFromColmnName( sheet, row, "Articulation Type" )
        group    = getGroupValue( sheet, row )

        # Must be required values to generate
        if( len( name ) == 0 or len( artiType ) == 0 ):
            continue

        print( "[Articulation] {name}, Type={type}, Group={group}".format(
            name = name,
            type = artiType,
            group = group
        ))

        artiType = constants.ARTICULATION_TYPE.index( artiType ) # to integer format ( 0: Attribute 1: Direction).

        ret += template.ARTICULATION.format(
            uuid1 = createUUID(),
            type  = artiType,
            name  = html.escape( name ),
            group = group
        )

    ret += template.ARTICULATION_FOOTER
    return ret

def genKeySwitch( sheet ):
    rowLength = sheet.max_row

    ret = template.KEY_SWITCH_HEADER.format(
        uuid1 = createUUID(),
        uuid2 = createUUID()
    )

    for row in range( 2, rowLength ):
        name  = xlsutil.getCellFromColmnName( sheet, row, "Name" )
        if len( name ) == 0:
            break
        articulation    = xlsutil.getCellFromColmnName( sheet, row, "Articulation" )
        color           = xlsutil.getCellFromColmnName( sheet, row, "Color" )
        group           = getGroupValue( sheet, row )
        noteNo          = xlsutil.getCellFromColmnName( sheet, row, "MIDI Note" )
        vel             = xlsutil.getCellFromColmnName( sheet, row, "Velocity" )
        ccNo            = xlsutil.getCellFromColmnName( sheet, row, "CC No." )
        ccValue         = xlsutil.getCellFromColmnName( sheet, row, "CC Value" )

        # float to int saferty
        vel     = float2int( vel )
        color   = float2int( color )
        ccNo    = float2int( ccNo, -1 )
        ccValue = float2int( ccValue, -1 )

        # Fail check
        if( len( name ) == 0 ):
            break

        print( "[Slot] {name}".format( name = name ) )

        # MIDI Message check( Note On )
        if( len( noteNo ) > 0 and noteNo in constants.NOTENUMBER ):
            noteNo = constants.NOTENUMBER.index( noteNo ) # to integer format (0-127)
        else:
            noteNo = -1
            vel    = 0

        # MIDI Message check( CC )
        if( ccNo < 0 or ccValue < 0 ):
            ccNo    = -1
            ccValue = -1

        # Append articulation
        if( len( articulation ) > 0 ):
            tmp = ""
            tmp += template.ARTICULATION_IN_SLOT_HEADER
            tmp += template.ARTICULATION_IN_SLOT.format(
                uuid1 = createUUID(), name = articulation,
                group = group
            )
            tmp += template.ARTICULATION_IN_SLOT_FOOTER
            articulation = tmp
        else:
            articulation = ""

        # Append MIDI message( Note On / CC )
        if( noteNo < 0 and ccNo < 0 ):
            midimessage = template.EMPTY_MIDI_MESSAGE_IN_KEYSWITCH
        else:
            midimessage = template.MIDI_MESSAGE_IN_KEYSWITCH_HEADER
            if( noteNo >= 0 ):
                midimessage += template.MIDI_MESSAGE_IN_KEYSWITCH.format(
                    uuid1 = createUUID(),
                    midiMessage = 144,
                    data1       = noteNo,
                    data2       = vel
                )
            if( ccNo >= 0 ):
                midimessage += template.MIDI_MESSAGE_IN_KEYSWITCH.format(
                    uuid1 = createUUID(),
                    midiMessage = 176,
                    data1       = ccNo,
                    data2       = ccValue
                )
            midimessage += template.MIDI_MESSAGE_IN_KEYSWITCH_FOOTER

        ret += template.KEY_SWITCH.format(
            uuid1 = createUUID(),
            uuid2 = createUUID(),
            uuid3 = createUUID(),
            uuid4 = createUUID(),
            midimessage = midimessage,
            name  = html.escape( name ),
            color = color,
            articulations = articulation
        )

    ret += template.SLOTS_FOOTER
    return ret

def usage():
    print(
"""---------------------
  XLS2ExpressionMap
  2017 (c) R-Koubou
---------------------
usage:
    python XLS2ExpressionMap.py <input file>
    <input file>: xlsx file path
        """
    )

def main():

    if( len( sys.argv ) < 2 ):
        usage()
        return

    xlsFilePath = sys.argv[ 1 ]
    book = openpyxl.load_workbook( filename = xlsFilePath, read_only = True )

    for sheetName in book.sheetnames:

        sheet = book[ sheetName ]

        # Ignore sheet
        if( sheetName == "DO NOT MODIFY!" ):
            continue

        expressionMapName = sheetName
        outputFileName    = expressionMapName + ".expressionmap"

        print( "Creating: {name}".format( name = outputFileName ) )
        fp = open( outputFileName, "wb" )

        fp.write( template.XML_HEADER.format( name = expressionMapName ).encode( "utf-8" ) )
        fp.write( genArticulation( sheet ).encode( "utf-8" ) )
        fp.write( genKeySwitch( sheet ).encode( "utf-8" ) )
        fp.write( template.XML_FOOTER.encode( "utf-8" ) )

        fp.close()

        print( "{name} created.".format(
            name = outputFileName
        ))
        print( "--------------------------------------------" )


if __name__ == '__main__':
    main()
