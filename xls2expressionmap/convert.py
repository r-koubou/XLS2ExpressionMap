# coding: utf-8

#-------------------
# built in modules
#-------------------
import sys
import html
from os import path

#-------------------
# 3rd party modules
#-------------------
# https://pypi.python.org/pypi/openpyxl
import openpyxl

#-------------------
# My Script
#-------------------
from . import constants
from .midi import mididata
from .util import util
from .expressionmap import template
from .xlsx import xlsutil

INT_MAX = sys.maxsize

"""
Convert to xlsx file to Cubase *.expressionmap.
"""
class XLS2ExpressionMap:

    """
    ctor
    """
    def __init__( self, xlsxFileName, outputDir = None ):
        self.xlsxFileName = xlsxFileName
        self.book         = openpyxl.load_workbook( filename = xlsxFileName, read_only = True )
        if outputDir != None:
            self.outputDir = outputDir
        else:
            self.outputDir = "."


    """
    Generate articulation xml string from given sheet
    """
    def generateArticulation( self, sheet, rows ):

        ret             = template.ARTICULATION_HEADER
        START_ROW_INEDX = xlsutil.getRowIndexCompat( rows )[ 1 ]

        for rowIndex in range( START_ROW_INEDX, sheet.max_row ):

            name = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_ARTICULATION )

            if name == None:
                break

            artiType = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_ARTICULATION_TYPE )
            group    = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_GROUP )

            # Must be required values to generate
            if name == None or artiType == None or group == None:
                continue

            # "group" xml value is zero origin
            group = group - 1

            print( "[Articulation] {name}, Type={type}, Group={group}".format(
                name  = name,
                type  = artiType,
                group = group
            ))

            artiType = constants.ARTICULATION_TYPE.index( artiType ) # to integer format ( 0: Attribute 1: Direction).

            ret += template.ARTICULATION.format(
                uuid1 = util.createUUID(),
                type  = artiType,
                name  = html.escape( name ),
                group = group
            )

        ret += template.ARTICULATION_FOOTER
        return ret

    """
    Generate velocity xml string from given data
    """
    def generateMIDINote( self, note ):

        return template.MIDI_MESSAGE_IN_KEYSWITCH.format(
                    uuid1       = util.createUUID(),
                    midiMessage = 144,
                    data1       = note.noteNo,
                    data2       = note.velocity
        )

    """
    Generate MIDI CC xml string from given data
    """
    def generateCC( self, cc ):

        return template.MIDI_MESSAGE_IN_KEYSWITCH.format(
                    uuid1       = util.createUUID(),
                    midiMessage = 176,
                    data1       = cc.ccNo,
                    data2       = cc.ccValue
        )

    """
    Generate MIDI Program change message xml string from given data
    """
    def generatePC( self, pc ):
        return template.MIDI_MESSAGE_IN_KEYSWITCH.format(
                    uuid1       = util.createUUID(),
                    midiMessage = 192,
                    data1       = pc.lsb,
                    data2       = pc.msb
        )

    """
    Generate key switch xml string from given sheet
    """
    def generateKeySwitch( self, sheet, rows ):

        ret             = template.KEY_SWITCH_HEADER
        START_ROW_INEDX = xlsutil.getRowIndexCompat( rows )[ 1 ]

        for rowIndex in range( START_ROW_INEDX, sheet.max_row ):

            name = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_NAME )
            if name == None:
                break

            articulation    = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_ARTICULATION )
            color           = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_COLOR )
            group           = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_GROUP )

            if group == None:
                group = 0
            else:
                # "group" xml value is zero origin
                group = group - 1

            print( "[Slot] {name}".format( name = name ) )

            # Append articulation
            if len( articulation ) > 0:
                artiType = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_ARTICULATION_TYPE )
                artiType = constants.ARTICULATION_TYPE.index( artiType ) # to integer format ( 0: Attribute 1: Direction).
                tmp = ""
                tmp += template.ARTICULATION_IN_SLOT_HEADER
                tmp += template.ARTICULATION_IN_SLOT.format(
                    uuid1               = util.createUUID(),
                    articulationtype    = artiType,
                    name = html.escape( articulation ),
                    group               = group,
                )
                tmp += template.ARTICULATION_IN_SLOT_FOOTER
                articulation = tmp
            else:
                articulation = ""

            # Append MIDI Notes
            # * Multiple MIDI Note Supported
            # * Column name format:
            #   MIDI Note1 ... MIDI Note1+n
            #   Velocity1 ... Velocity1+n
            midiNoteList = []
            for i in range( 1, INT_MAX ):
                noteNo = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_MIDI_NOTE + str( i ) )
                vel    = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_MIDI_VELOCITY + str( i ) )
                if noteNo == None or vel == None:
                    break

                if noteNo in constants.NOTENUMBER:
                    noteNo = constants.NOTENUMBER.index( noteNo ) # to integer format (0-127)

                obj = mididata.MIDINote( noteNo, vel )
                if obj.valid():
                    midiNoteList.append( obj )

            # Append MIDI CC
            # * Multiple MIDI CC Supported
            # * Column name format:
            #   CC No1 ... CC No1+n
            #   CC Value1 ... CC Value1+n
            ccList = []
            for i in range( 1, INT_MAX ):
                ccNo    = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_MIDI_CC + str( i ) )
                ccValue = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_MIDI_CC_VALUE + str( i ) )
                if ccNo == None or ccValue == None:
                    break

                obj = mididata.MIDICc( ccNo, ccValue )
                if obj.valid():
                    ccList.append( obj )

            # Append MIDI PC
            # * Multiple MIDI Program Change Supported
            # * Column name format:
            #   PC LSB1 ... PC LSB1+n
            #   PC MSB1 ... PC MSB1+n (MSB not exist, MSB value will be 0 )
            programChangeList = []
            for i in range( 1, INT_MAX ):
                lsb = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_MIDI_PC_LSB + str( i ) )
                msb = xlsutil.getValueFromColumnName( rows, rowIndex, constants.COLUMN_MIDI_PC_MSB + str( i ) )

                if lsb == None:
                    break
                elif msb == None:
                    # msb is not set
                    msb = 0

                obj = mididata.MIDIPc( lsb, msb )
                if obj.valid():
                    programChangeList.append( obj )

            # Generate MIDI message xml node
            midiMessageXml = ""
            if len( midiNoteList ) > 0 or len( ccList ) > 0 or len( programChangeList ) > 0:

                midiMessageXml += template.MIDI_MESSAGE_IN_KEYSWITCH_HEADER

                # MIDI Note
                for n in midiNoteList:
                    midiMessageXml += self.generateMIDINote( n )

                # MIDI CC
                for cc in ccList:
                    midiMessageXml  += self.generateCC( cc )

                # MIDI PC
                for pc in programChangeList:
                    midiMessageXml += self.generatePC( pc )

                midiMessageXml += template.MIDI_MESSAGE_IN_KEYSWITCH_FOOTER
            else:
                midiMessageXml = template.EMPTY_MIDI_MESSAGE_IN_KEYSWITCH

            # Append key xml node
            #
            #                <int name="key{i}" value="{note}"/>
            #
            keyXml = ""
            if len( midiNoteList ) == 0:
                keyXml = "               <int name=\"key\" value=\"-1\"/>"
            else:
                for index, i in enumerate( midiNoteList ):
                    if index >= 1:
                        keyXml += template.KEY_SWITCH_NOTE_VALUE.format(
                            keyindex    = str( index - 1 ),
                            note        = i.noteNo
                        )
                    else:
                        keyXml += template.KEY_SWITCH_NOTE_VALUE.format(
                            keyindex    = "",
                            note        = i.noteNo
                        )

                    keyXml += "\n"

            # Generate Keyswitch XML text
            ret += template.KEY_SWITCH.format(
                uuid1 = util.createUUID(),
                uuid2 = util.createUUID(),
                uuid3 = util.createUUID(),
                uuid4 = util.createUUID(),
                midimessage = midiMessageXml,
                name  = html.escape( name ),
                color = color,
                keynote_values = keyXml,
                articulations = articulation
            )

        ret += template.SLOTS_FOOTER
        return ret

    """
    Convert to Expression map file
    """
    def convert( self ):
        try:
            for sheetName in self.book.sheetnames:

                xmlText = ''
                sheet   = self.book[ sheetName ]
                rows    = [x for x in sheet.rows] # Tuple [row][column]
                # Ignore sheet
                if sheetName == constants.LIST_DEFINITION_SHEETNAME:
                    continue

                # Since: 0.7
                # Sheet Name's Max length is 32...
                # If A1 is "Output Name" and B1 is <Valid File Name>
                # outputFileName will be assigned instead of sheet's name

                # V0.6.x mode
                expressionMapName = sheetName
                outputFileName    = expressionMapName + ".expressionmap"
                outputFileName    = path.join( self.outputDir, outputFileName )

                # V0.7 later mode
                a0 = rows[ 0 ][ 0 ].value
                b0 = rows[ 0 ][ 1 ].value
                if a0 == constants.COLUMN_OUTPUTNAME:
                    b0 = str( b0 )
                    expressionMapName = b0
                    outputFileName    = expressionMapName + ".expressionmap"
                    outputFileName    = path.join( self.outputDir, outputFileName )
                # Switching to V0.6 mode
                else:
                    print( "----------------------------------------------------------------" )
                    print( "Run as V0.6 mode" )
                    print( "Warning: Please use spreadsheet version 0.7 or later format." )
                    print( "In the future, this compatibility will be removed." )
                    print( "----------------------------------------------------------------" )

                xmlText  = template.XML_HEADER.format( name = expressionMapName )
                xmlText += self.generateArticulation( sheet, rows )
                xmlText += self.generateKeySwitch( sheet, rows )
                xmlText += template.XML_FOOTER

                xmlText = xmlText.encode( 'utf8' )

                fp = open( outputFileName, "wb" )
                fp.write( xmlText )
                fp.close()
        finally:
            self.book.close()
